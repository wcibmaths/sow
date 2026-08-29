from flask import Flask, request, jsonify, send_file, abort
import os
import psycopg2
import psycopg2.extras
from psycopg2.pool import ThreadedConnectionPool
from openpyxl import load_workbook

app = Flask(__name__)

_pool = None

def get_pool():
    global _pool
    if _pool is None:
        _pool = ThreadedConnectionPool(
            minconn=1,
            maxconn=5,
            dsn=os.environ['DATABASE_URL']
        )
    return _pool

def load_data():
    pool = get_pool()
    conn = pool.getconn()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("SELECT data FROM sow_store WHERE id = 1")
            row = cur.fetchone()
            return dict(row['data']) if row else {}
    finally:
        pool.putconn(conn)

def save_data(data):
    pool = get_pool()
    conn = pool.getconn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO sow_store (id, data, updated_at)
                VALUES (1, %s::jsonb, NOW())
                ON CONFLICT (id) DO UPDATE
                    SET data = EXCLUDED.data,
                        updated_at = NOW()
                """,
                (psycopg2.extras.Json(data),)
            )
        conn.commit()
    finally:
        pool.putconn(conn)

@app.route('/')
def index():
    return send_file('index.html')

@app.route('/wellington-logo.png')
def wellington_logo():
    return send_file('attached_assets/wellington-logo.png', mimetype='image/png')

@app.route('/exam-dates.js')
def exam_dates():
    return send_file('exam-dates.js', mimetype='application/javascript')

@app.route('/y10-fm-sow.js')
def y10_fm_sow_source():
    return send_file('y10-fm-sow.js', mimetype='application/javascript')

@app.route('/api/data', methods=['GET'])
def get_data():
    return jsonify(load_data())

@app.route('/api/data', methods=['POST'])
def post_data():
    data = request.get_json(force=True, silent=True)
    if data is None:
        abort(400)
    save_data(data)
    return jsonify({'ok': True})

@app.route('/api/y10-accelerated-sow', methods=['GET'])
def get_y10_accelerated_sow():
    """Return the Y10 Accelerated sheet in the app's SoW row shape."""
    workbook_path = os.path.join(
        os.path.dirname(__file__),
        'attached_assets',
        'KS4_Accelerated_SoW_(3)_1787970692790.xlsx',
    )
    try:
        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        sheet = workbook['Y10 Accelerated']
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or '').strip() for value in rows[1]]
        records = []

        def calendar_term(dates):
            """Place untitled workbook calendar rows in the term they begin."""
            text = str(dates or '').lower()
            import re
            match = re.search(
                r'\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)',
                text,
            )
            first_month = match.group(1) if match else ''
            if first_month in ('jan', 'feb', 'mar'):
                return 'Lent'
            if first_month in ('apr', 'may', 'jun', 'jul'):
                return 'Summer'
            return 'Michaelmas'

        for index, values in enumerate(rows[2:], start=1):
            source = {
                header: (value if value is not None else '')
                for header, value in zip(headers, values)
            }
            cycle = str(source.get('Cycle') or '').strip()
            week = str(source.get('Week') or '').strip()
            if not cycle and week:
                cycle = week[-1]
            unit = source.get('Unit')
            if unit == '':
                unit = None
            objective = str(source.get('Objective') or '').strip()
            specification = str(source.get('Specification point(s)') or '').strip()
            qualification = str(source.get('Qualification') or '').strip()
            pages = str(source.get('Student Book pages') or '').strip()
            personalisation = str(
                source.get('Personalisation and Stretch') or ''
            ).strip()
            lesson_type = str(source.get('Type') or '').strip() or 'core'
            term = str(source.get('Term') or '').strip()
            if not term and lesson_type == 'calendar':
                term = calendar_term(source.get('Dates'))
            detail_parts = [
                f'Objective: {objective}' if objective else '',
                f'Specification: {specification}'
                if specification and specification != '—' else '',
                f'Qualification: {qualification}'
                if qualification and qualification != '—' else '',
                f'Student Book: {pages}' if pages and pages != '—' else '',
                f'Personalisation and stretch: {personalisation}'
                if personalisation else '',
            ]
            row = {
                'id': f'y10_fm_source_{index:03d}',
                'week': week,
                'dates': str(source.get('Dates') or '').strip(),
                'cycle': cycle,
                'unit': unit,
                'unitName': str(source.get('Unit Name') or '').strip(),
                'lessonNo': str(source.get('Lesson No.') or '').strip(),
                'lessonName': str(source.get('Lesson Name') or '').strip(),
                'type': lesson_type,
                'term': term,
                'obj': '  |  '.join(part for part in detail_parts if part),
                'objective': objective,
                'specification': specification,
                'qualification': qualification,
                'pages': pages,
                'personalisation': personalisation,
                'alert': str(source.get('Alert / Notes') or '').strip(),
            }
            records.append(row)
        return jsonify(records)
    except Exception as exc:
        app.logger.exception('Unable to load Y10 Accelerated sheet')
        return jsonify({'error': str(exc)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False)
